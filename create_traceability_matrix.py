import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Sheet 1: User Creation Flows
ws1 = wb.active
ws1.title = 'User Creation Flows'

# Headers
headers1 = ['Endpoint', 'HTTP Method', 'Controller', 'Service Method', 'Audit Events', 'Control IDs', 'Status']
for col, header in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Data
flows = [
    ['/api/auth/register', 'POST', 'AccountApiController', 'IAuthenticationService.RegisterAsync', 'AM01_USER_CREATED, AM01_USER_REGISTERED', 'AM-01', 'Implemented'],
    ['/api/trial/signup', 'POST', 'TrialApiController', 'ITrialLifecycleService.SignUpForTrialAsync', 'AM01_TRIAL_SIGNUP_INITIATED', 'AM-01, AM-09', 'Implemented'],
    ['/api/trial/provision', 'POST', 'TrialApiController', 'ITrialLifecycleService.ProvisionTrialAsync', 'AM01_TENANT_CREATED, AM01_USER_CREATED, AM03_ROLE_ASSIGNED', 'AM-01, AM-03, AM-09', 'Implemented'],
    ['/api/tenants/{tenantId}/users/invite', 'POST', 'UserInvitationController', 'IUserInvitationService.InviteUserAsync', 'AM01_USER_INVITED', 'AM-01', 'Implemented'],
    ['/api/invitation/accept', 'POST', 'InvitationAcceptController', 'IUserInvitationService.AcceptInvitationAsync', 'AM01_USER_CREATED, AM03_ROLE_ASSIGNED', 'AM-01, AM-03', 'Implemented'],
]

for row_idx, row_data in enumerate(flows, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws1.cell(row=row_idx, column=col_idx, value=value)

# Adjust column widths
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 25
ws1.column_dimensions['D'].width = 45
ws1.column_dimensions['E'].width = 55
ws1.column_dimensions['F'].width = 20
ws1.column_dimensions['G'].width = 15

# Sheet 2: Control to Code Mapping
ws2 = wb.create_sheet('Control to Code')

headers2 = ['Control ID', 'Control Name', 'Implementation File', 'Key Classes/Methods', 'Audit Events', 'Status']
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

controls = [
    ['AM-01', 'User Identity Management', 'Controllers/AccountApiController.cs', 'Register, Login', 'AM01_USER_CREATED, AM01_USER_REGISTERED, AM01_USER_INVITED', 'Implemented'],
    ['AM-02', 'Authentication Security', 'Services/AuthenticationService.Identity.cs', 'LoginAsync, ValidatePassword', 'AM02_LOGIN_SUCCESS, AM02_LOGIN_FAILED, AM02_MFA_VERIFIED', 'Implemented'],
    ['AM-03', 'Role-Based Access Control', 'Configuration/RoleConstants.cs', 'RoleConstants, AssignRole', 'AM03_ROLE_ASSIGNED, AM03_ROLE_REMOVED, AM03_PERMISSION_GRANTED', 'Implemented'],
    ['AM-04', 'Access Reviews', 'Controllers/Api/AccessReviewController.cs', 'AccessReviewService (PLANNED)', 'AM04_REVIEW_CREATED, AM04_DECISION_RECORDED', 'In Development'],
    ['AM-05', 'Privileged Access Management', 'Services/PrivilegedAccessService.cs', 'ValidatePrivilegedAccess', 'AM05_PRIV_ACCESS_GRANTED, AM05_PRIV_ACCESS_USED', 'Implemented'],
    ['AM-06', 'Service Account Management', 'Services/ServiceAccountService.cs', 'CreateServiceAccount', 'AM06_SERVICE_ACCOUNT_CREATED', 'Implemented'],
    ['AM-07', 'Session Management', 'Services/SessionService.cs', 'CreateSession, ValidateSession', 'AM07_SESSION_CREATED, AM07_SESSION_TERMINATED', 'Implemented'],
    ['AM-08', 'Password Policy', 'Configuration/PasswordPolicyOptions.cs', 'ValidatePassword', 'AM08_PASSWORD_CHANGED, AM08_PASSWORD_EXPIRED', 'Implemented'],
    ['AM-09', 'Trial Lifecycle', 'Services/TrialLifecycleService.cs', 'ProvisionTrialAsync, ExpireTrial', 'AM09_TRIAL_STARTED, AM09_TRIAL_EXPIRED', 'Implemented'],
    ['AM-10', 'Audit Logging', 'Services/AuditEventService.cs', 'LogEventAsync', 'AM10_AUDIT_EVENT_LOGGED, AM10_AUDIT_EXPORTED', 'Implemented'],
    ['AM-11', 'Periodic Access Reviews', 'Jobs/AccessReviewReminderJob.cs', 'AccessReviewCampaignService (PLANNED)', 'AM11_REVIEW_SCHEDULED, AM11_REVIEW_COMPLETED', 'In Development'],
    ['AM-12', 'Segregation of Duties', 'Services/SoDService.cs (PLANNED)', 'ValidateSoD (PLANNED)', 'AM12_SOD_VIOLATION_DETECTED', 'Planned'],
    ['MG-01', 'Module Inventory', 'Configuration/ModuleGovernanceControls.cs', 'AbpModuleRegister', 'MG01_MODULE_ADDED, MG01_INVENTORY_REVIEWED', 'Implemented'],
    ['MG-02', 'Module Change Approval', 'Configuration/ModuleGovernanceControls.cs', 'MG02_ChangeApproval', 'MG02_CHANGE_APPROVED, MG02_CHANGE_DEPLOYED', 'Implemented'],
    ['MG-03', 'Environment Parity', 'Configuration/ModuleGovernanceControls.cs', 'MG03_EnvironmentParity', 'MG03_PARITY_CHECK_PASSED', 'Implemented'],
    ['AU-01', 'Access Logging', 'Services/AuditEventService.cs', 'LogAccessEvent', 'AU01_LOGIN_LOGGED, AU01_LOGOUT_LOGGED', 'Implemented'],
    ['AU-02', 'Business Event Logging', 'Services/AuditEventService.cs', 'LogBusinessEvent', 'AU02_EVENT_CREATED, AU02_EVENT_APPROVED', 'Implemented'],
    ['AU-03', 'Platform Admin Audit', 'Services/AuditEventService.cs', 'LogPlatformEvent', 'AU03_ADMIN_ACTION_LOGGED', 'Implemented'],
    ['FM-01', 'Feature Flag Governance', 'Configuration/FeatureManagementControls.cs', 'GrcFeatures', 'FM01_FEATURE_ENABLED, FM01_FEATURE_DISABLED', 'Implemented'],
    ['BP-01', 'Background Job Governance', 'Configuration/BackgroundProcessingControls.cs', 'HangfireJobRegistry', 'BP01_JOB_COMPLETED, BP01_JOB_FAILED', 'Implemented'],
    ['BP-02', 'ABP Worker Disablement', 'Configuration/BackgroundProcessingControls.cs', 'BP02_WorkerDisablement', 'BP02_DISABLEMENT_DOCUMENTED', 'Documented'],
    ['IN-01', 'Integration Approval', 'Configuration/IntegrationControls.cs', 'IN01_IntegrationApproval', 'IN01_INTEGRATION_ENABLED', 'Implemented'],
    ['IN-02', 'Credential Management', 'Configuration/IntegrationControls.cs', 'IN02_CredentialManagement', 'IN02_CREDENTIAL_ROTATED', 'Implemented'],
    ['AI-01', 'AI Feature Enablement', 'Configuration/AIGovernanceControls.cs', 'AI01_FeatureEnablement', 'AI01_OPTIN_APPROVED, AI01_FEATURE_ENABLED', 'Implemented'],
    ['AI-02', 'AI Usage Logging', 'Configuration/AIGovernanceControls.cs', 'AI02_UsageLogging', 'AI02_REQUEST_COMPLETED', 'Implemented'],
]

for row_idx, row_data in enumerate(controls, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_data[-1] == 'In Development':
            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        elif row_data[-1] == 'Planned':
            cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 45
ws2.column_dimensions['D'].width = 40
ws2.column_dimensions['E'].width = 50
ws2.column_dimensions['F'].width = 15

# Sheet 3: Audit Events Summary
ws3 = wb.create_sheet('Audit Events')

headers3 = ['Event Type', 'Control ID', 'Description', 'Severity', 'Retention']
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

events = [
    ['AM01_USER_CREATED', 'AM-01', 'New user account created', 'Info', '7 years'],
    ['AM01_USER_REGISTERED', 'AM-01', 'User self-registered', 'Info', '7 years'],
    ['AM01_USER_INVITED', 'AM-01', 'User invited by admin', 'Info', '7 years'],
    ['AM01_TENANT_CREATED', 'AM-01', 'New tenant provisioned', 'Info', '7 years'],
    ['AM01_TRIAL_SIGNUP_INITIATED', 'AM-01', 'Trial signup started', 'Info', '7 years'],
    ['AM02_LOGIN_SUCCESS', 'AM-02', 'Successful authentication', 'Info', '7 years'],
    ['AM02_LOGIN_FAILED', 'AM-02', 'Failed authentication attempt', 'Warning', '7 years'],
    ['AM02_ACCOUNT_LOCKED', 'AM-02', 'Account locked due to failures', 'Warning', '7 years'],
    ['AM02_MFA_VERIFIED', 'AM-02', 'MFA verification successful', 'Info', '7 years'],
    ['AM03_ROLE_ASSIGNED', 'AM-03', 'Role assigned to user', 'Info', '7 years'],
    ['AM03_ROLE_REMOVED', 'AM-03', 'Role removed from user', 'Info', '7 years'],
    ['AM03_PERMISSION_GRANTED', 'AM-03', 'Permission granted', 'Info', '7 years'],
    ['AM04_REVIEW_CREATED', 'AM-04', 'Access review initiated', 'Info', '7 years'],
    ['AM04_DECISION_RECORDED', 'AM-04', 'Review decision recorded', 'Info', '7 years'],
    ['AM04_REVIEW_COMPLETED', 'AM-04', 'Access review completed', 'Info', '7 years'],
    ['AM05_PRIV_ACCESS_GRANTED', 'AM-05', 'Privileged access granted', 'Warning', '7 years'],
    ['AM05_PRIV_ACCESS_USED', 'AM-05', 'Privileged access exercised', 'Warning', '7 years'],
    ['AM09_TRIAL_STARTED', 'AM-09', 'Trial period started', 'Info', '7 years'],
    ['AM09_TRIAL_EXPIRED', 'AM-09', 'Trial period expired', 'Info', '7 years'],
    ['AM09_TRIAL_CONVERTED', 'AM-09', 'Trial converted to paid', 'Info', '7 years'],
]

for row_idx, row_data in enumerate(events, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws3.cell(row=row_idx, column=col_idx, value=value)

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 40
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 12

# Sheet 4: Implementation Status
ws4 = wb.create_sheet('Implementation Status')

headers4 = ['Control Family', 'Total Controls', 'Implemented', 'In Development', 'Planned', 'Coverage %']
for col, header in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

status = [
    ['Access Management (AM)', 12, 9, 2, 1, '75%'],
    ['Module Governance (MG)', 3, 3, 0, 0, '100%'],
    ['Auditability (AU)', 3, 3, 0, 0, '100%'],
    ['Feature Management (FM)', 1, 1, 0, 0, '100%'],
    ['Background Processing (BP)', 2, 2, 0, 0, '100%'],
    ['Integration (IN)', 2, 2, 0, 0, '100%'],
    ['AI Governance (AI)', 2, 2, 0, 0, '100%'],
    ['TOTAL', 25, 22, 2, 1, '88%'],
]

for row_idx, row_data in enumerate(status, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == len(status) + 1:
            cell.font = Font(bold=True)

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 15
ws4.column_dimensions['C'].width = 15
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 12

wb.save('Control_to_Code_Traceability_Matrix.xlsx')
print('Created Control_to_Code_Traceability_Matrix.xlsx successfully')
